import csv
from faker import Faker
from openpyxl import Workbook
from datetime import datetime

# Ініціалізація Faker для української мови
fake = Faker('uk_UA')

# Функція для генерації персональних даних та збереження їх у CSV
def generate_personal_data(filename, num_records=2000):
    with open(filename, mode='w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        writer.writerow(
            ['Прізвище', 'Ім’я', 'По батькові', 'Стать', 'Дата народження', 'Посада', 'Місто', 'Адреса', 'Телефон', 'Email']
        )
        for _ in range(num_records):
            gender = fake.random_element(elements=('male', 'female'))
            last_name = fake.last_name_male() if gender == 'male' else fake.last_name_female()
            first_name = fake.first_name_male() if gender == 'male' else fake.first_name_female()
            patronymic = fake.middle_name_male() if gender == 'male' else fake.middle_name_female()
            birth_date = fake.date_of_birth(minimum_age=16, maximum_age=85)
            position = fake.job()
            city = fake.city()
            address = fake.address()
            phone = fake.phone_number()
            email = fake.email()

            writer.writerow([last_name, first_name, patronymic, gender, birth_date, position, city, address, phone, email])

# Функція для зчитування CSV і створення XLSX файлу з розподілом за віковими категоріями
def create_xlsx_from_csv(csv_filename, xlsx_filename):
    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "all"
    ws_younger_18 = wb.create_sheet("younger_18")
    ws_18_45 = wb.create_sheet("18-45")
    ws_45_70 = wb.create_sheet("45-70")
    ws_older_70 = wb.create_sheet("older_70")

    with open(csv_filename, mode='r', encoding='utf-8') as file:
        reader = csv.reader(file)
        headers = next(reader)

        # Додавання заголовків для кожного аркуша
        for ws in [ws_all, ws_younger_18, ws_18_45, ws_45_70, ws_older_70]:
            ws.append(headers + ['Вік'])

        # Зчитування рядків і розподіл за віковими категоріями
        for row in reader:
            birth_date = datetime.strptime(row[4], '%Y-%m-%d')
            age = datetime.today().year - birth_date.year - (
                (datetime.today().month, datetime.today().day) < (birth_date.month, birth_date.day)
            )
            row.append(str(age))

            # Додавання рядків до всіх аркушів
            ws_all.append(row)

            # Розподіл по вікових категоріях
            if age < 18:
                ws_younger_18.append(row)
            elif 18 <= age <= 45:
                ws_18_45.append(row)
            elif 45 < age <= 70:
                ws_45_70.append(row)
            else:
                ws_older_70.append(row)

    # Збереження файлу Excel
    try:
        wb.save(xlsx_filename)
        print("Файл XLSX успішно створено.")
    except Exception as e:
        print(f"Помилка створення XLSX файлу: {e}")

# Виклик функцій для генерації даних і створення Excel
generate_personal_data('personal_data.csv')
create_xlsx_from_csv('personal_data.csv', 'personal_data.xlsx')
